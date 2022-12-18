# projectForScott.py
# A program to support SEO, take search criteria from excel spreadsheet, convert to HTML scripts
# Created by Katie S. Johnston, INFOST 350-201, 12-01-2022

# first, import openpyxl
import openpyxl

# next, give location of file
path = "testExcel.xlsx"
try:
    # Open workbook by creating object
    wb_object = openpyxl.load_workbook(path)

    # Get workbook active sheet object
    sheet_object = wb_object.active

    # Now cell object test to get specific cell and row...
    cell_object = sheet_object.cell(row=2, column=4)

    # super complex, but basically I am splitting to get the specific state for each URL
    # This will be crucial for getting the keywords to correlate to the correct string
    split_item_test = cell_object.value.split(
        "https://www.trustedchoice.com/l/")
    split_item_test_2 = split_item_test[1].split("/")
    split_item_test_3 = split_item_test_2[0]

    # Getting the value of maximum rows
    # and column
    row = sheet_object.max_row
    column = sheet_object.max_column
    print("Total Rows:", row)
    print("Total Columns:", column)

    print("\nValue of fourth column")
    state_dict = {}
    for row in sheet_object.iter_rows(min_row=2, max_row=sheet_object.max_row):
        k = row[3].value.split(
            'https://www.trustedchoice.com/l/')[1].split("/")[0].strip()
        # future state - If there's more than 10 results per state, only print the top 10 results. Determined by rank numbers in the file...
        r = row[1].value
        v = row[0].value
        if k in state_dict:
            state_dict[k] += [v]
        else:
            state_dict[k] = [v]

    # Print values - this is to verify what it looks like before entering in the text file
    for key, value in state_dict.items():
        print(f'{key}, Total {len(value)}', end='; ')
        for v in value:
            print(f'{v}', end=', ')
        print('')

    # write the values to a text file, if the file doesn't exist python will make one
    with open('resulting_states.txt', 'w') as results:
        # Print values
        for key, value in state_dict.items():
            results.write(f"State: {key}=")
            for v in value:
                results.write(f"{v},")
            results.write(";\n \n")

    print("All good")
except FileNotFoundError:
	msg = "Sorry, the file " + path + " does not exist"
	print(msg)
